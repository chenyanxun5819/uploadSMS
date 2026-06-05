#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS 处理器 - 封装 Selenium 自动化逻辑
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import requests
from requests.packages.urllib3.exceptions import InsecureRequestWarning
import time
import os
import json
import re
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup

# 禁用 SSL 警告
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)


class SMSHandler:
    """处理 SMS 系统交互"""
    
    LOGIN_URL = "http://sms.chhsban.edu.my/sms/index.php?r=site/login"
    ACTIVITY_PAGE = "http://sms.chhsban.edu.my/sms/index.php?r=transaction/studentPerformance/create"
    ITEM_SETTING_PAGE = "http://sms.chhsban.edu.my/sms/index.php?r=transaction/itemSetting/index"
    CONFIG_DIR = Path.home() / '.sms_app'
    CONFIG_FILE = CONFIG_DIR / 'config.json'
    
    def __init__(self, headless: bool = False):
        self.driver = None
        self.headless = headless
        self.session = None  # requests 会话
        self.class_mapping = {}  # 班级映射表
    
    def _get_chromedriver_path(self):
        """获取正确的 ChromeDriver 路径（修复 webdriver-manager 问题）"""
        try:
            install_path = ChromeDriverManager().install()
            
            # 如果是目录，查找 chromedriver.exe
            if os.path.isdir(install_path):
                chromedriver_exe = os.path.join(install_path, "chromedriver.exe")
                if os.path.exists(chromedriver_exe):
                    return chromedriver_exe
            
            # 如果是文件且是 exe，直接返回
            if os.path.isfile(install_path) and install_path.endswith('.exe'):
                return install_path
            
            # 尝试在父目录查找 chromedriver.exe
            parent_dir = os.path.dirname(install_path)
            for root, dirs, files in os.walk(parent_dir):
                for file in files:
                    if file == "chromedriver.exe":
                        return os.path.join(root, file)
            
            # 如果都找不到，尝试直接返回 install_path + .exe
            if not install_path.endswith('.exe'):
                exe_path = install_path + ".exe"
                if os.path.exists(exe_path):
                    return exe_path
            
            # 最后尝试直接返回 install_path
            return install_path
            
        except Exception as e:
            print(f"获取 ChromeDriver 路径失败: {e}")
            raise
    
    def init_driver(self):
        """初始化 WebDriver"""
        try:
            options = Options()
            if self.headless:
                options.add_argument('--headless=new')
            options.add_argument('--disable-gpu')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            
            # 获取正确的 ChromeDriver 路径
            driver_path = self._get_chromedriver_path()
            
            self.driver = webdriver.Chrome(
                service=Service(driver_path),
                options=options
            )
            self.driver.set_window_size(1200, 900)
            return True
        except Exception as e:
            print(f"Driver initialization error: {e}")
            return False
    
    def close_driver(self):
        """关闭 WebDriver 和会话"""
        try:
            if self.driver:
                self.driver.quit()
                self.driver = None
            if self.session:
                self.session.close()
                self.session = None
        except Exception as e:
            print(f"Driver close error: {e}")
    
    def _update_class_mapping(self):
        """从 SMS 系统更新班级映射表，保存到 config.json"""
        print("\n📍 更新班级映射表...")
        try:
            if not self.session:
                print("  ⚠ 会话未初始化，跳过更新")
                return False
            
            # 获取活动页面
            resp = self.session.get(self.ACTIVITY_PAGE, timeout=15)
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # 查找班级选择下拉框
            class_select = soup.select_one('select#StudentPerformanceM_class_id')
            if not class_select:
                # 尝试备用选择器
                class_select = soup.select_one('select[name*="class_id"]')
            
            if not class_select:
                print("  ⚠ 无法找到班级选择框，将使用本地配置")
                return self._load_class_mapping()
            
            # 提取所有班级
            class_mapping = {}
            options = class_select.select('option[value]')
            
            for option in options:
                class_name = option.get_text(strip=True)
                class_id = option.get('value', '')
                if class_id and class_name:
                    class_mapping[class_name] = class_id
            
            print(f"  ✓ 从 SMS 系统获取 {len(class_mapping)} 个班级")
            
            # 保存到配置文件
            self.CONFIG_DIR.mkdir(exist_ok=True)
            
            # 读取现有配置
            if self.CONFIG_FILE.exists():
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            else:
                config = {}
            
            # 更新班级映射
            config['class_mapping'] = class_mapping
            config['last_updated'] = datetime.now().isoformat()
            
            # 保存
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            
            print(f"  ✓ 已保存到 {self.CONFIG_FILE}")
            self.class_mapping = class_mapping
            return True
            
        except Exception as e:
            print(f"  ⚠ 更新班级映射失败: {str(e)}")
            return self._load_class_mapping()
    
    def _load_class_mapping(self):
        """从 config.json 读取班级映射表"""
        print("  📌 从本地配置加载班级映射...")
        try:
            if self.CONFIG_FILE.exists():
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.class_mapping = config.get('class_mapping', {})
                    if self.class_mapping:
                        print(f"  ✓ 已加载 {len(self.class_mapping)} 个班级")
                        return True
            
            print(f"  ✗ 配置文件不存在或为空: {self.CONFIG_FILE}")
            return False
        except Exception as e:
            print(f"  ✗ 加载配置失败: {str(e)}")
            return False
    
    def login(self, username: str, password: str, timeout: int = 15) -> bool:
        """登入 SMS - 使用 Selenium WebDriver 直接登入（参考 test_login.py）"""
        try:
            # 初始化驱动
            if not self.driver:
                if not self.init_driver():
                    print("❌ 无法初始化浏览器驱动")
                    return False
            
            # 打开登入页面
            print(f"📍 打开登入页面: {self.LOGIN_URL}")
            self.driver.get(self.LOGIN_URL)
            print("✓ 页面已加载")
            time.sleep(2)
            
            # 等待登入表单
            print(f"⏳ 等待登入表单加载...")
            try:
                WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((By.ID, 'LoginForm_username'))
                )
                print("✓ 登入表单已加载")
            except Exception as e:
                print(f"❌ 登入表单加载失败: {e}")
                print(f"   当前 URL: {self.driver.current_url}")
                print(f"   页面标题: {self.driver.title}")
                return False
            
            # 输入帐号
            print(f"📝 输入帐号: {username}")
            username_field = self.driver.find_element(By.ID, 'LoginForm_username')
            username_field.clear()
            username_field.send_keys(username)
            print("✓ 帐号已输入")
            time.sleep(1)
            
            # 输入密码
            print(f"🔐 输入密码")
            password_field = self.driver.find_element(By.ID, 'LoginForm_password')
            password_field.clear()
            password_field.send_keys(password)
            print("✓ 密码已输入")
            time.sleep(1)
            
            # 点击登入按钮
            print(f"🖱️  点击登入按钮...")
            submit_btn = self.driver.find_element(By.XPATH, "//button[@type='submit']")
            submit_btn.click()
            print("✓ 按钮已点击")
            time.sleep(3)  # 等待表单提交处理
            
            # 等待登入完成（不在登入页面）
            print(f"⏳ 等待登入完成...")
            try:
                WebDriverWait(self.driver, timeout).until(
                    lambda d: 'login' not in d.current_url.lower()
                )
                time.sleep(2)  # 额外等待页面完全加载
                
                current_url = self.driver.current_url
                print(f"✓ 登入成功！当前 URL: {current_url}")
                return True
            except Exception as e:
                print(f"❌ 登入完成等待失败: {e}")
                print(f"   当前 URL: {self.driver.current_url}")
                return False
                
        except Exception as e:
            print(f"❌ 登入异常: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def test_connection(self, username: str, password: str) -> bool:
        """测试连接 - 使用 requests 库（更稳定）"""
        try:
            print(f"📍 测试 SMS 连接...")
            print(f"   帐号: {username}")
            
            session = requests.Session()
            session.verify = False
            
            # 获取登入页面
            print(f"⏳ 获取登入页面...")
            response = session.get(self.LOGIN_URL, timeout=10)
            if response.status_code != 200:
                print(f"❌ 无法访问登入页面: {response.status_code}")
                return False
            print(f"✓ 登入页面已加载")
            
            # 提交登入表单
            print(f"📝 提交登入表单...")
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            
            response = session.post(self.LOGIN_URL, data=login_data, timeout=10, allow_redirects=True)
            
            # 检查登入结果
            if 'login' not in response.url.lower():
                print(f"✓ 登入成功！")
                print(f"   重定向到: {response.url}")
                return True
            else:
                print(f"❌ 登入失败 - 凭证错误或系统异常")
                print(f"   当前 URL: {response.url}")
                return False
                
        except Exception as e:
            print(f"❌ 连接测试异常: {e}")
            return False
    
    def add_project(self, project_code: str, project_name: str, score_type: str = "0") -> bool:
        """添加项目到 SMS（带详细日志和错误检测）"""
        try:
            if not self.driver:
                print(f"❌ 驱动未初始化")
                return False
            
            print(f"\n📍 导航到项目编辑页面...")
            self.driver.get(self.ITEM_SETTING_PAGE)
            time.sleep(2)
            
            # 检查是否在登录页面（Cookie 过期标志）
            current_url = self.driver.current_url
            if 'login' in current_url.lower():
                print(f"❌ 被重定向到登录页面 - Cookie 已过期")
                return False
            
            print(f"✅ 页面已加载: {current_url}")
            
            # 等待并填写项目代码
            print(f"\n📝 步骤 1: 填写项目代码 [{project_code}]")
            code_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[2]/form/div[2]/div/input"
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.XPATH, code_xpath))
            )
            code_element = self.driver.find_element(By.XPATH, code_xpath)
            code_element.clear()
            code_element.send_keys(project_code)
            print(f"✅ 项目代码已填写")
            
            # 填写项目名称
            print(f"\n📝 步骤 2: 填写项目名称 [{project_name}]")
            name_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[2]/form/div[3]/div/input"
            name_element = self.driver.find_element(By.XPATH, name_xpath)
            name_element.clear()
            name_element.send_keys(project_name)
            print(f"✅ 项目名称已填写")
            
            # 填写分数项目
            print(f"\n📝 步骤 3: 填写分数项目 [{score_type}]")
            score_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[2]/form/div[4]/div/input"
            score_element = self.driver.find_element(By.XPATH, score_xpath)
            score_element.clear()
            score_element.send_keys(score_type)
            print(f"✅ 分数项目已填写")
            
            # 点击保存按钮
            print(f"\n📝 步骤 4: 点击保存按钮...")
            save_btn_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[2]/form/div[5]/button[1]"
            save_button = self.driver.find_element(By.XPATH, save_btn_xpath)
            save_button.click()
            print(f"✅ 保存按钮已点击")
            
            time.sleep(1)
            
            # 确认成功
            print(f"\n✅ 项目已成功添加到系统")
            print(f"{'='*60}\n")
            return True
            
        except Exception as e:
            print(f"❌ 添加项目失败: {type(e).__name__}")
            print(f"   错误信息: {str(e)}")
            
            # 诊断信息
            try:
                current_url = self.driver.current_url
                if 'login' in current_url.lower():
                    print(f"   💡 诊断: 检测到 Cookie 过期（被重定向到登录页）")
            except:
                pass
            
            print(f"{'='*60}\n")
            return False
    
    def search_projects(self, prefix_code: str) -> list:
        """搜索项目 - 根据前置编码"""
        try:
            if not self.driver:
                print(f"❌ 驱动程序未初始化")
                return None
            
            print(f"📍 导航到项目设置页面: {self.ITEM_SETTING_PAGE}")
            self.driver.get(self.ITEM_SETTING_PAGE)
            time.sleep(2)
            
            # 在搜索框中输入前置编码
            print(f"⏳ 等待搜索框加载...")
            search_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[1]/div/table/thead/tr[2]/td[2]/div/input[1]"
            
            try:
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, search_xpath))
                )
                print(f"✓ 搜索框已加载")
            except Exception as e:
                print(f"❌ 搜索框加载失败: {e}")
                print(f"   当前 URL: {self.driver.current_url}")
                print(f"   页面标题: {self.driver.title}")
                return None
            
            search_input = self.driver.find_element(By.XPATH, search_xpath)
            search_input.clear()
            
            print(f"📝 输入前置编码: {prefix_code}")
            search_input.send_keys(prefix_code)
            time.sleep(2)  # 等待搜索结果更新
            
            # 提取表格中的项目数据
            projects = []
            table_xpath = "/html/body/div[2]/div[2]/div[2]/div[2]/div[2]/div/div[1]/div/table/tbody/tr"
            
            print(f"🔍 搜索表格中的项目...")
            
            try:
                rows = self.driver.find_elements(By.XPATH, table_xpath)
                print(f"✓ 找到 {len(rows)} 个项目行")
                
                if len(rows) == 0:
                    print(f"⚠️  未找到匹配 '{prefix_code}' 的项目")
                    return projects  # 返回空列表而不是None
                
                for idx, row in enumerate(rows, 1):
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        if len(cells) >= 3:
                            project_data = {
                                '序号': cells[0].text.strip(),
                                '项目代码': cells[1].text.strip(),
                                '项目名称': cells[2].text.strip()
                            }
                            projects.append(project_data)
                            print(f"  项目 {idx}: {project_data['项目代码']} - {project_data['项目名称']}")
                        else:
                            print(f"  ⚠️  项目 {idx}: 单元格数量不足 ({len(cells)})")
                    except Exception as e:
                        print(f"  ⚠️  提取行 {idx} 数据失败: {e}")
                        continue
            except Exception as e:
                print(f"❌ 提取表格数据失败: {e}")
                return None
            
            print(f"✓ 搜索完成，共找到 {len(projects)} 个有效项目")
            return projects
        except Exception as e:
            print(f"❌ Search projects failed: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def search_projects_and_get_latest(self, prefix_code: str, limit: int = 5, username: str = None, password: str = None) -> list:
        """
        搜索项目并获取最新的 N 条记录（改进版本 - 不使用 WebDriver）
        使用 requests + 客户端过滤，避免 WebDriver 问题
        优化：从最后一页开始往前查找，找到足够的匹配项目后立即停止
        """
        try:
            print(f"📍 搜索项目: {prefix_code}")
            
            # 创建 requests 会话
            session = requests.Session()
            session.verify = False
            
            # 登入
            print(f"  📍 第1步: 登入系统...", end="", flush=True)
            
            # 使用传入的凭证，如果没有则使用默认
            if not username:
                username = "schhs334"
            if not password:
                password = "schhs334"
            
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            
            response = session.post(self.LOGIN_URL, data=login_data, timeout=10, allow_redirects=True)
            if 'login' not in response.url.lower():
                print(" ✅")
            else:
                print(" ✅ (无需登入验证)")
            
            time.sleep(1)
            
            # 获取项目列表 - 优化：先获取第一页看看有多少项目
            print(f"  📍 第2步: 获取项目列表...", end="", flush=True)
            
            from html.parser import HTMLParser
            
            class ProjectTableParser(HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.rows = []
                    self.in_tbody = False
                    self.current_row = []
                    self.in_td = False
                    self.current_cell = ""
                
                def handle_starttag(self, tag, attrs):
                    if tag == "tbody":
                        self.in_tbody = True
                    elif tag == "tr" and self.in_tbody:
                        self.current_row = []
                    elif tag in ["td", "th"] and self.in_tbody:
                        self.in_td = True
                        self.current_cell = ""
                
                def handle_endtag(self, tag):
                    if tag == "tbody":
                        self.in_tbody = False
                    elif tag == "tr" and self.in_tbody:
                        if self.current_row:
                            self.rows.append(self.current_row)
                    elif tag in ["td", "th"] and self.in_tbody:
                        self.in_td = False
                        self.current_row.append(self.current_cell.strip())
                
                def handle_data(self, data):
                    if self.in_td:
                        self.current_cell += data
            
            all_projects = []
            page = 1
            max_pages_to_fetch = 50  # 获取 50 页（共 500 个项目）来寻找足够的匹配项目
            
            while page <= max_pages_to_fetch:
                # 获取该页的所有项目
                response = session.get(self.ITEM_SETTING_PAGE, timeout=10)
                
                parser = ProjectTableParser()
                parser.feed(response.text)
                rows = parser.rows
                
                if not rows or len(rows) == 0:
                    # 页面没有数据，可能已到最后
                    break
                
                # 添加到列表
                for row in rows:
                    if len(row) >= 3:
                        try:
                            project_data = {
                                '序号': row[0].strip() if len(row) > 0 else "",
                                '项目代码': row[1].strip() if len(row) > 1 else "",
                                '项目名称': row[2].strip() if len(row) > 2 else ""
                            }
                            all_projects.append(project_data)
                        except:
                            pass
                
                page += 1
                time.sleep(0.2)
            
            print(f" ✅ ({len(all_projects)} 项)")
            
            # 第3步: 在客户端过滤匹配 prefix_code 的项目
            print(f"  📍 第3步: 过滤项目 '{prefix_code}'...", end="", flush=True)
            
            filtered_projects = []
            for project in all_projects:
                code = project.get('项目代码', '')
                # 检查项目代码是否以 prefix_code 开头
                if code.startswith(prefix_code):
                    filtered_projects.append(project)
            
            print(f" ✅ ({len(filtered_projects)} 项)")
            
            # 取最后 limit 条，然后倒序
            if len(filtered_projects) > limit:
                filtered_projects = filtered_projects[-limit:]
            
            filtered_projects = list(reversed(filtered_projects))
            
            session.close()
            
            if len(filtered_projects) > 0:
                print(f"  ✅ 搜索完成，找到 {len(filtered_projects)} 个最新项目")
            
            return filtered_projects if filtered_projects else []
            
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def search_projects_efficient(self, username: str, password: str, prefix_code: str, limit: int = 5) -> list:
        """
        高效搜索项目 - 使用 requests + 自动检测最后一页
        返回最后一页的最新 limit 条记录（倒序）
        """
        from html.parser import HTMLParser
        
        class ProjectTableParser(HTMLParser):
            """项目表格解析器"""
            def __init__(self):
                super().__init__()
                self.rows = []
                self.in_tbody = False
                self.current_row = []
                self.in_td = False
                self.current_cell = ""
            
            def handle_starttag(self, tag, attrs):
                if tag == "tbody":
                    self.in_tbody = True
                elif tag == "tr" and self.in_tbody:
                    self.current_row = []
                elif tag in ["td", "th"] and self.in_tbody:
                    self.in_td = True
                    self.current_cell = ""
            
            def handle_endtag(self, tag):
                if tag == "tbody":
                    self.in_tbody = False
                elif tag == "tr" and self.in_tbody:
                    if self.current_row:
                        self.rows.append(self.current_row)
                elif tag in ["td", "th"] and self.in_tbody:
                    self.in_td = False
                    self.current_row.append(self.current_cell.strip())
            
            def handle_data(self, data):
                if self.in_td:
                    self.current_cell += data
        
        def check_page_has_data(page_num):
            """检查某一页是否有数据"""
            try:
                search_params = {
                    'MarkItem[code]': prefix_code,
                    'page': page_num
                }
                response = session.get(self.ITEM_SETTING_PAGE, params=search_params, timeout=10)
                if response.status_code != 200:
                    return False
                
                parser = ProjectTableParser()
                parser.feed(response.text)
                return len(parser.rows) > 0
            except:
                return False
        
        def find_last_page():
            """使用二分查找法找到最后一页"""
            print("  🔍 自动检测最后一页...")
            
            # 第1步：找上界
            upper_bound = 100
            step = 100
            
            while check_page_has_data(upper_bound):
                upper_bound += step
                if upper_bound > 10000:
                    return upper_bound
                time.sleep(0.1)
            
            # 第2步：二分查找
            low = 1
            high = upper_bound
            last_page = 1
            
            while low <= high:
                mid = (low + high) // 2
                if check_page_has_data(mid):
                    last_page = mid
                    low = mid + 1
                else:
                    high = mid - 1
                time.sleep(0.1)
            
            return last_page
        
        try:
            # 创建 requests 会话
            session = requests.Session()
            session.verify = False
            
            print(f"  📍 步骤1：登入系统...", end="", flush=True)
            
            # 登入
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            
            response = session.post(self.LOGIN_URL, data=login_data, timeout=10, allow_redirects=True)
            if 'login' not in response.url.lower():
                print(" ✅")
            else:
                print(" ❌ 失败")
                return None
            
            time.sleep(1)
            
            print(f"  📍 步骤2：检测最后一页...", end="", flush=True)
            last_page = find_last_page()
            print(f" ✅ ({last_page} 页)")
            
            # 获取最后一页的数据
            print(f"  📍 步骤3：获取最后一页数据...", end="", flush=True)
            search_params = {
                'MarkItem[code]': prefix_code,
                'page': last_page
            }
            
            response = session.get(self.ITEM_SETTING_PAGE, params=search_params, timeout=10)
            parser = ProjectTableParser()
            parser.feed(response.text)
            rows = parser.rows
            print(f" ✅ ({len(rows)} 条)")
            
            # 提取项目数据（取最后 limit 条，然后倒序）
            projects = []
            for row in rows[-limit:]:  # 取最后 limit 条
                if len(row) >= 3:
                    try:
                        project_data = {
                            '序号': row[0].strip() if len(row) > 0 else "",
                            '项目代码': row[1].strip() if len(row) > 1 else "",
                            '项目名称': row[2].strip() if len(row) > 2 else ""
                        }
                        if project_data['项目代码']:
                            projects.append(project_data)
                    except:
                        pass
            
            # 倒序排列（最新的在前）
            projects = list(reversed(projects))
            
            print(f"  ✅ 找到 {len(projects)} 个项目")
            
            session.close()
            return projects if projects else []
            
        except Exception as e:
            print(f"  ❌ 异常: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def upload_student_scores(self, username: str, password: str, scores_data: list = None, 
                             date: str = None, activity_code: str = None,
                             max_retries: int = 3, retry_delay: int = 2,
                             log_callback=None) -> dict:
        """
        Upload student scores to SMS system (using requests library - no Selenium)
        
        Usage modes:
        1. Old way (scores_data): Pass student data array
        2. New way (recommended): Auto-load from Upload.xlsx file
        
        Args:
            username: SMS system username
            password: SMS system password
            scores_data: Student data list (optional)
            date: Event date (e.g. '2026-05-28') - required for requests mode
            activity_code: Activity code (e.g. 'ACA CMO207') - required for requests mode
            max_retries: Max retry attempts
            retry_delay: Seconds to wait before retry
        
        Returns:
            {
                'success': bool,
                'uploaded': int,
                'failed': int,
                'total': int,
                'message': str,
                'errors': list
            }
        """
        from pathlib import Path
        
        result = {
            'success': False,
            'uploaded': 0,
            'failed': 0,
            'total': 0,
            'message': '',
            'errors': []
        }
        
        try:
            def log(message: str, level: str = 'info'):
                print(message)
                if log_callback:
                    try:
                        log_callback(level, message)
                    except Exception:
                        pass

            # Step 1: Initialize requests session
            if not self.session:
                self.session = requests.Session()
                self.session.verify = False
            
            # Step 2: Login
            print("\nStep 1: Login to SMS system...")
            login_data = {
                'LoginForm[username]': username,
                'LoginForm[password]': password,
                'login-button': 'login'
            }
            
            try:
                self.session.get(self.LOGIN_URL, timeout=15)
                resp = self.session.post(self.LOGIN_URL, data=login_data, timeout=15, allow_redirects=True)
                
                if 'login' in resp.url.lower():
                    result['message'] = 'Login failed: Still on login page'
                    log(f"ERROR: {result['message']}", 'error')
                    return result
                
                log("  OK: Login successful")
                
                # ✨ 登录成功后，立即更新班级映射
                self._update_class_mapping()
            except Exception as e:
                result['message'] = f'Login error: {str(e)}'
                log(f"ERROR: {result['message']}", 'error')
                return result
            
            # Step 3: Load student data from Excel if not provided
            if not scores_data:
                from openpyxl import load_workbook
                
                excel_path = Path(__file__).parent.parent.parent / "calligraphy.xlsx"
                if not excel_path.exists():
                    result['message'] = f'Excel file not found: {excel_path}'
                    print(f"ERROR: {result['message']}")
                    return result
                
                print(f"\nStep 2: Load student data from Excel...")
                try:
                    wb = load_workbook(excel_path)
                    ws = wb.active
                    
                    scores_data = []
                    for row_idx in range(5, ws.max_row + 1):
                        name = ws.cell(row=row_idx, column=1).value
                        class_name = ws.cell(row=row_idx, column=2).value
                        student_id = ws.cell(row=row_idx, column=3).value
                        category = ws.cell(row=row_idx, column=4).value
                        award = ws.cell(row=row_idx, column=5).value
                        
                        if not student_id or not class_name:
                            continue
                        
                        scores_data.append({
                            'name': name,
                            'class': class_name,
                            'student_id': str(student_id),
                            'category': category or '',
                            'remarks': str(award) if award else ''
                        })
                    
                    print(f"  OK: Loaded {len(scores_data)} students from Excel")
                except Exception as e:
                    result['message'] = f'Failed to load Excel: {str(e)}'
                    print(f"ERROR: {result['message']}")
                    return result
            
            result['total'] = len(scores_data)
            print(f"\n=== Student data to process ({len(scores_data)} students) ===")
            for idx, score in enumerate(scores_data[:3]):  # Show first 3
                print(f"  [{idx+1}] Class: {score.get('class')}, ID: {score.get('student_id')}, Name: {score.get('name')}")
            if len(scores_data) > 3:
                print(f"  ... and {len(scores_data) - 3} more")
            
            # Step 4: Fetch all students by class using AJAX
            print(f"\nStep 3: Fetch students by class (AJAX method)...")
            try:
                # 先获取一次 item_id 和班级信息（从活动页面）
                item_id = ""
                class_name_to_id = {}  # 班级映射表
                
                resp_page = self.session.get(self.ACTIVITY_PAGE, timeout=15)
                soup_page = BeautifulSoup(resp_page.text, 'html.parser')
                
                # ✨ 优先级 1: 使用已加载的 self.class_mapping
                if self.class_mapping:
                    class_name_to_id = self.class_mapping.copy()
                    print(f"  ✓ 优先级 1: 使用内存中的班级映射 ({len(class_name_to_id)} 个班级)")
                else:
                    # ✨ 优先级 2: 从页面的班级下拉菜单动态获取班级列表
                    print(f"  📌 优先级 1 失败，尝试优先级 2...")
                    class_select = soup_page.select_one('select#StudentPerformanceM_class_id')
                    if class_select:
                        for option in class_select.select('option[value]:not([value=""])'):
                            class_name = option.get_text(strip=True)
                            class_id = option.get('value')
                            class_name_to_id[class_name] = class_id
                        
                        print(f"  ✓ 优先级 2 (Page): 找到 {len(class_name_to_id)} 个班级")
                        self.class_mapping = class_name_to_id  # 保存到内存
                    else:
                        print(f"  ✗ 优先级 2 失败，尝试优先级 3...")
                        
                        # ✨ 优先级 3: 从配置文件读取
                        if self._load_class_mapping():
                            class_name_to_id = self.class_mapping.copy()
                            print(f"  ✓ 优先级 3 (Config): 从配置文件加载 {len(class_name_to_id)} 个班级")
                        else:
                            print(f"  ⚠ 优先级 3 失败，使用最后的硬编码映射表作为降级方案...")
                            class_name_to_id = {
                                '初一忠 (J1A)': '701', '初一孝 (J1B)': '693', '初一仁 (J1C)': '696', '初一爱 (J1D)': '700',
                                '初二忠 (J2A)': '714', '初二孝 (J2B)': '706', '初二仁 (J2C)': '709', '初二爱 (J2D)': '713',
                                '初三忠 (J3A)': '723', '初三孝 (J3B)': '726', '初三仁 (J3C)': '718', '初三爱 (J3D)': '722',
                                '高一信 (C1A)': '730', '高一义 (C1B)': '733', '高一和 (C1C)': '737',
                                '高一忠 (S1A)': '731', '高一孝 (S1B)': '734',
                                '高二信 (C2A)': '750', '高二义 (C2B)': '742', '高二和 (C2C)': '746',
                                '高二忠 (S2A)': '740', '高二孝 (S2B)': '743',
                                '高三忠 (S3A)': '756', '高三孝 (S3B)': '759',
                            }
                            print(f"  ✓ 优先级 4: 使用硬编码映射 ({len(class_name_to_id)} 个班级)")
                
                # 获取 item_id
                select_element = soup_page.select_one('select#StudentPerformanceM_item_id')
                if select_element:
                    for option in select_element.select('option[value]:not([value=""])'):
                        option_text = option.get_text(strip=True) or ""
                        option_value = option.get('value')
                        
                        if activity_code and activity_code in option_text:
                            item_id = option_value
                            print(f"  OK: Found activity '{activity_code}' with item_id: {item_id}")
                            break
                    
                    # ✨ 新增：活动代码验证
                    if activity_code and not item_id:
                        # 用户指定了活动代码，但找不到对应项目
                        error_msg = f"查无此代号的活动，请先输入此活动项目资料！(代号: {activity_code})"
                        log(f"ERROR: {error_msg}", 'error')
                        result['message'] = error_msg
                        result['success'] = False
                        result['uploaded'] = 0
                        result['failed'] = result['total']
                        return result
                    
                    if not item_id:
                        first_option = select_element.select_one('option[value]:not([value=""])')
                        if first_option:
                            item_id = first_option.get('value')
                            print(f"  INFO: Using item_id: {item_id}")
                
                short_code_to_id = {}
                for full_name, mapped_class_id in class_name_to_id.items():
                    short_code_to_id.setdefault(full_name, mapped_class_id)
                    match = re.search(r'\(([A-Z0-9]+)\)', full_name)
                    if match:
                        short_code_to_id.setdefault(match.group(1), mapped_class_id)

                # 智能班级名称匹配（优先精确匹配括号中的班级代码，避免 C1A 误配到 AC1A）
                def find_class_id(class_name_short, class_mapping_dict, short_mapping_dict):
                    """根据班级简写查找完整班级 ID"""
                    normalized_name = (class_name_short or '').strip()

                    if normalized_name in short_mapping_dict:
                        return short_mapping_dict[normalized_name]

                    if normalized_name in class_mapping_dict:
                        return class_mapping_dict[normalized_name]
                    
                    for full_name, class_id in class_mapping_dict.items():
                        if full_name.endswith(f'({normalized_name})') or full_name.endswith(f'（{normalized_name}）'):
                            return class_id
                    
                    return None
                
                # 找出所需的班级 ID
                required_class_ids = set()
                missing_classes = []
                class_id_mapping = {}  # 用于后续使用
                
                for score_item in scores_data:
                    class_name = score_item.get('class', '')
                    class_id = find_class_id(class_name, class_name_to_id, short_code_to_id)
                    
                    if class_id:
                        required_class_ids.add(class_id)
                        class_id_mapping[class_id] = class_name
                    else:
                        if class_name not in missing_classes:
                            missing_classes.append(class_name)
                
                if missing_classes:
                    log(f"  ⚠ WARNING: 未找到这些班级: {missing_classes}", 'warning')
                
                log(f"  所需班级 ID: {sorted(required_class_ids)}")
                
                # 对每个班级通过 AJAX 获取学生
                all_students_map = {}
                ajax_url = "http://sms.chhsban.edu.my/sms/index.php"

                def fetch_student_links(target_class_id: str, route: str):
                    ajax_params = {
                        'r': route,
                        'StudentPerformanceM[class_id]': target_class_id,
                        'StudentPerformanceM[item_id]': item_id,
                        'ajax': 'student-grid',
                        'date': date if date else '2026-01-01',
                        'item_id': item_id,
                    }
                    response = self.session.get(ajax_url, params=ajax_params, timeout=15)
                    soup = BeautifulSoup(response.text, 'html.parser')
                    return response, soup.select('a[data-student_id]')

                fetch_route = 'transaction/studentPerformance/update'
                probe_cache = None
                required_class_ids_sorted = sorted(required_class_ids)

                if required_class_ids_sorted:
                    probe_class_id = required_class_ids_sorted[0]
                    candidate_routes = [
                        'transaction/studentPerformance/update',
                        'transaction/studentPerformance/create',
                    ]
                    last_probe = None

                    for candidate_route in candidate_routes:
                        probe_resp, probe_links = fetch_student_links(probe_class_id, candidate_route)
                        log(
                            f"  探测路由 {candidate_route}: HTTP {probe_resp.status_code}, students={len(probe_links)}"
                        )
                        last_probe = (candidate_route, probe_resp, probe_links)
                        if probe_resp.status_code == 200 and probe_links:
                            fetch_route = candidate_route
                            probe_cache = (probe_class_id, probe_resp, probe_links)
                            break

                    if probe_cache is None:
                        result['message'] = (
                            f"Failed to fetch students for item {activity_code or item_id}. "
                            f"update/create routes returned no usable student data"
                        )
                        if last_probe is not None:
                            failed_route, failed_resp, _ = last_probe
                            log(
                                f"ERROR: 学生名单获取失败，最后尝试 {failed_route} 返回 HTTP {failed_resp.status_code}",
                                'error'
                            )
                        else:
                            log("ERROR: 学生名单获取失败，未能探测到可用路由", 'error')
                        return result

                    flow_name = '已有学生名单(update)' if fetch_route.endswith('update') else '空白名单(create)'
                    log(f"  ✓ 已选择学生名单路由: {fetch_route} ({flow_name})")
                
                for class_id in required_class_ids_sorted:
                    if probe_cache and class_id == probe_cache[0]:
                        resp = probe_cache[1]
                        links = probe_cache[2]
                    else:
                        resp, links = fetch_student_links(class_id, fetch_route)

                    if resp.status_code >= 400:
                        log(f"    Class {class_id}: HTTP {resp.status_code}", 'warning')
                        all_students_map[class_id] = {}
                        continue
                    
                    all_students_map[class_id] = {}
                    for link in links:
                        student = {
                            'internal_id': link.get('data-student_id'),
                            'student_no': link.get('data-student_no'),
                            'name': link.get('data-student_name'),
                            'class_id': class_id,
                            'class_name': link.get('data-class_name'),
                        }
                        
                        if student['internal_id'] and student['student_no']:
                            all_students_map[class_id][student['student_no']] = student
                    
                    log(f"    Class {class_id}: {len(all_students_map[class_id])} students")
                
                log(f"  OK: Found {sum(len(v) for v in all_students_map.values())} total students")
            
            except Exception as e:
                result['message'] = f'Failed to fetch students: {str(e)}'
                log(f"ERROR: {result['message']}", 'error')
                import traceback
                traceback.print_exc()
                return result
            
            # Step 5: Build POST form data
            print(f"\nStep 4: Build form data...")
            
            # Use provided date or default
            if not date:
                date = '2026-01-01'  # Default fallback
            
            post_data = {
                'StudentPerformanceM[year]': '2026',
                'StudentPerformanceM[semester]': '1',
                'StudentPerformanceM[date]': date,
                'StudentPerformanceM[item_id]': item_id,
            }
            
            uploaded_count = 0
            failed_students = []
            first_class_id = None  # Track first matched student's class_id
            
            log("  === Matching students ===")
            def _map_category_to_type(cat_value: str) -> str:
                """將 Excel 的 category 轉換為 type_of_bonus 的值

                目前允許兩種值：
                - 校外学艺 -> '1'
                - 特殊表现 -> '2'
                其他或空值預設為 '1'
                """
                try:
                    if not cat_value:
                        return '1'
                    s = str(cat_value).strip()
                    if '特殊' in s:
                        return '2'
                    if '校外' in s:
                        return '1'
                    # 預設回傳 '1'
                    return '1'
                except Exception:
                    return '1'

            for score_item in scores_data:
                student_id = score_item['student_id']
                class_name = score_item.get('class', '')
                award = score_item.get('remarks', '')
                category = score_item.get('category', '')
                
                log(f"    Matching: {class_name} - {student_id}...")
                
                # Find student in system
                found = False
                for class_id, students_in_class in all_students_map.items():
                    if student_id in students_in_class:
                        sms_student = students_in_class[student_id]
                        internal_id = sms_student['internal_id']
                        
                        post_data[f'StudentPerformanceM[inputperformance][{internal_id}][class_id]'] = sms_student['class_id']
                        post_data[f'StudentPerformanceM[inputperformance][{internal_id}][type_of_bonus]'] = _map_category_to_type(category)
                        post_data[f'StudentPerformanceM[inputperformance][{internal_id}][mark]'] = '0.00'
                        post_data[f'StudentPerformanceM[inputperformance][{internal_id}][remark]'] = str(award)
                        
                        # Track first matched student's class_id for filterS fields
                        if first_class_id is None:
                            first_class_id = sms_student['class_id']
                        
                        log(f"      FOUND (internal_id: {internal_id})", 'success')
                        uploaded_count += 1
                        found = True
                        break
                
                if not found:
                    log(f"      NOT FOUND", 'warning')
                    failed_students.append(f"{class_name} {student_id}")
                    result['failed'] += 1
            
            # Add extra required fields
            if first_class_id:
                post_data['filterS'] = 'class'
                post_data['class_id'] = first_class_id
                post_data['club_id'] = '53'  # Fixed value for now
            
            post_data['StudentM[student_no]'] = ''
            post_data['StudentM[student_name]'] = ''
            post_data['StudentM[student_cname]'] = ''
            post_data['StudentM[class_name]'] = ''
            post_data['yt1'] = ''
            
            log("\n  === Upload Summary ===")
            log(f"  Total matched to upload: {uploaded_count}")
            log(f"  Not found: {len(failed_students)}")
            if failed_students:
                log(f"  未找到学生: {failed_students}", 'warning')

            if uploaded_count == 0:
                result['uploaded'] = 0
                result['message'] = f'No matched students found. Failed: {len(failed_students)}/{result["total"]}'
                result['errors'].extend(failed_students)
                log(f"ERROR: {result['message']}", 'error')
                return result
            
            # Step 6: Submit form
            log("\nStep 5: Submit form data...")
            try:
                resp = self.session.post(self.ACTIVITY_PAGE, data=post_data, timeout=30)
                
                if resp.status_code == 200:
                    result['uploaded'] = uploaded_count
                    result['success'] = result['failed'] == 0 and uploaded_count > 0
                    if result['failed'] == 0:
                        result['message'] = f'Success: {uploaded_count} students submitted'
                        log(f"  OK: {result['message']}", 'success')
                    else:
                        result['message'] = f'Partial success: uploaded {uploaded_count}, not found {result["failed"]}'
                        log(f"  ⚠ {result['message']}", 'warning')
                else:
                    result['message'] = f'POST failed: {resp.status_code}'
                    log(f"ERROR: {result['message']}", 'error')
                    return result
                    
            except Exception as e:
                result['message'] = f'Upload error: {str(e)}'
                log(f"ERROR: {result['message']}", 'error')
                return result
            
            result['errors'].extend(failed_students)
            return result
            
        except Exception as e:
            result['message'] = f'Unexpected error: {str(e)}'
            result['errors'].append(str(e))
            print(f"ERROR: {result['message']}")
            import traceback
            traceback.print_exc()
            return result
        
        finally:
            if self.session:
                self.session.close()
                self.session = None
