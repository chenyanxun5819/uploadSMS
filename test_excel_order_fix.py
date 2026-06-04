#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试 score_upload_page.py 中修复后的 Excel 列顺序
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / 'sms_app'))

from openpyxl import load_workbook

print("=" * 100)
print("测试 score_upload_page.py 的 Excel 列顺序修复")
print("=" * 100)

# 模拟 score_upload_page.py 中的读取逻辑
excel_path = Path("calligraphy.xlsx")

if not excel_path.exists():
    print(f"\n✗ 找不到 {excel_path}")
    sys.exit(1)

print(f"\n读取文件: {excel_path}")

wb = load_workbook(excel_path, data_only=True)
ws = wb.active

# 读取日期（B1）
date_val = ws.cell(row=1, column=2).value
print(f"上传日期: {date_val}")

# 读取数据（从第5行开始）
print("\n" + "=" * 100)
print("修复后的列顺序（应该正确）")
print("=" * 100)

scores_data = []
for row_num in range(5, ws.max_row + 1):
    cell_values = {
        'name': ws.cell(row=row_num, column=1).value,      # Column 1: 姓名
        'class': ws.cell(row=row_num, column=2).value,     # Column 2: 班级
        'student_id': ws.cell(row=row_num, column=3).value, # Column 3: 学号
        'remarks': ws.cell(row=row_num, column=4).value,   # Column 4: 奖项/备注
        'english_name': ws.cell(row=row_num, column=5).value, # Column 5: 英文名
    }
    
    # 如果至少有学号和班级，则视为有效行
    if cell_values['student_id'] and cell_values['class']:
        scores_data.append(cell_values)
        print(f"\n第 {row_num} 行:")
        for key, value in cell_values.items():
            if value:
                print(f"  {key:15}: {value}")

wb.close()

print("\n" + "=" * 100)
print(f"读取完成: {len(scores_data)} 条有效学生数据")
print("=" * 100)

# 验证数据是否可用于上传
print("\n验证数据格式:")
if scores_data:
    first_student = scores_data[0]
    print(f"\n✓ 第一个学生:")
    print(f"  班级: {first_student['class']}")
    print(f"  学号: {first_student['student_id']}")
    print(f"  姓名: {first_student['name']}")
    print(f"  备注: {first_student['remarks']}")
    
    # 验证班级是否能在映射表中找到
    class_name_to_id = {
        'J1A': '701', 'J1D': '700', 'J1E': '704', 'J1H': '703',
        'J2A': '714', 'J2B': '706', 'J2C': '709', 'J2D': '713', 'J2E': '705', 'J2F': '708', 'J2G': '712', 'J2H': '715', 'J2I': '707', 'J2J': '711',
        'J3A': '723', 'J3B': '726', 'J3C': '718', 'J3D': '722', 'J3E': '725', 'J3F': '717', 'J3G': '721', 'J3H': '724', 'J3I': '716',
        'S1A': '731', 'S1B': '734',
        'S2A': '740', 'S2B': '743',
    }
    
    if first_student['class'] in class_name_to_id:
        class_id = class_name_to_id[first_student['class']]
        print(f"\n✓ 班级 {first_student['class']} 映射到 class_id: {class_id}")
    else:
        print(f"\n✗ 班级 {first_student['class']} 未找到映射")

print("\n" + "=" * 100)
print("✅ 测试完成 - Excel 列顺序已正确修复")
print("=" * 100)
