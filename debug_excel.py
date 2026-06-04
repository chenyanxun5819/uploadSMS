#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试脚本 - 检查 post_data 的构建
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# 查看 Upload.xlsx 的内容
print("=" * 80)
print("检查 Upload.xlsx 内容")
print("=" * 80)

excel_path = Path(__file__).parent / "calligraphy.xlsx"

if excel_path.exists():
    print(f"\n✓ 文件找到: {excel_path}")
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    print(f"\n工作表名: {ws.title}")
    print(f"最大行数: {ws.max_row}")
    print(f"最大列数: {ws.max_column}")
    
    print(f"\n=== 前几行内容 ===")
    for row_idx in range(1, min(10, ws.max_row + 1)):
        values = []
        for col_idx in range(1, min(6, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            values.append(str(cell.value) if cell.value else "")
        print(f"Row {row_idx}: {' | '.join(values)}")
    
    print(f"\n=== 学生数据 (从第5行开始) ===")
    scores_data = []
    for row_idx in range(5, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        class_name = ws.cell(row=row_idx, column=2).value
        student_id = ws.cell(row=row_idx, column=3).value
        award = ws.cell(row=row_idx, column=4).value
        
        if not student_id or not class_name:
            continue
        
        scores_data.append({
            'name': name,
            'class': class_name,
            'student_id': str(student_id),
            'remarks': str(award) if award else ''
        })
    
    print(f"找到 {len(scores_data)} 条学生记录:")
    for idx, score in enumerate(scores_data, 1):
        print(f"  [{idx}] Class: {score['class']:10} | Student ID: {score['student_id']:10} | Name: {score['name']:15} | Award: {score['remarks']}")
    
    print("\n✓ Excel 文件读取成功")
else:
    print(f"\n✗ 文件未找到: {excel_path}")

print("\n" + "=" * 80)
