"""
完整测试脚本：登入 SMS 系统，使用 Select2 操作活动下拉菜单，抓取活动名称，写入 Excel B3
"""

import time
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook

# 配置
SMS_URL = "http://sms.chhsban.edu.my/sms/index.php?r=site/login"
USERNAME = "schhs334"
PASSWORD = "schhs334"
ACTIVITY_PAGE = "http://sms.chhsban.edu.my/sms/index.php?r=transaction/studentPerformance/create"
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Upload.xlsx")

def login_to_sms(driver):
    """登入 SMS 系统"""
    print("[1] 连接到登入页面...")
    driver.get(SMS_URL)
    time.sleep(2)
    
    print("[2] 填入凭证...")
    driver.find_element(By.ID, "LoginForm_username").send_keys(USERNAME)
    driver.find_element(By.ID, "LoginForm_password").send_keys(PASSWORD)
    
    print("[3] 提交表单...")
    driver.find_element(By.XPATH, "//button[@type='submit']").click()
    
    # 等待重定向离开登入页面
    print("[4] 等待登入完成...")
    WebDriverWait(driver, 10).until(
        lambda d: "login" not in d.current_url.lower()
    )
    time.sleep(2)
    print("✓ 登入成功！")

def get_activity_name(driver, activity_code: str):
    """
    导航到活动页面并获取活动名称（使用 Select2 API + 直接 value 设置）
    
    Args:
        driver: Selenium WebDriver
        activity_code: 活动代码（从 Excel A2 读取）
    
    Returns:
        str: 活动名称，若失败返回 "查無此活動"
    """
    print(f"\n[5] 导航到活动页面...")
    driver.get(ACTIVITY_PAGE)
    time.sleep(3)
    
    print(f"[6] 查找活动代码: {activity_code}")
    
    try:
        # 方法 1: 通过 Select2 搜索 (使用 jQuery 的 select2 API)
        print("    方法 1: 通过 Select2 搜索...")
        
        search_script = f"""
        var select = $('#StudentPerformanceM_item_id');
        
        // 打开 Select2 并搜索
        select.select2('open');
        var searchInput = document.querySelector('.select2-input');
        if (searchInput) {{
            searchInput.value = '{activity_code}';
            // 触发搜索事件
            searchInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
            searchInput.dispatchEvent(new Event('keyup', {{ bubbles: true }}));
            searchInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
        }}
        
        return 'searching...';
        """
        
        driver.execute_script(search_script)
        time.sleep(2)
        
        # 获取搜索结果
        result_script = """
        var resultLabel = document.querySelector('.select2-result-label');
        if (resultLabel && resultLabel.innerText.trim()) {
            return resultLabel.innerText;
        }
        return null;
        """
        
        activity_name = driver.execute_script(result_script)
        
        if activity_name and activity_name.strip():
            print(f"    发现活动: {activity_name}")
            
            # 点击选中
            click_script = """
            var resultLabel = document.querySelector('.select2-result-label');
            if (resultLabel) {
                resultLabel.click();
                return true;
            }
            return false;
            """
            
            clicked = driver.execute_script(click_script)
            if clicked:
                time.sleep(1)
                print(f"✓ 成功选择活动: {activity_name}")
                return activity_name
        
        print("    方法 1 失败，尝试方法 2...")
        
        # 方法 2: 直接通过 hidden select 的 value 属性设置
        print("    方法 2: 直接设置 hidden select value...")
        
        set_value_script = f"""
        var select = $('#StudentPerformanceM_item_id');
        // 获取所有 option 并找出匹配的 value
        var options = select.find('option');
        var targetValue = null;
        
        for (var i = 0; i < options.length; i++) {{
            var text = options[i].innerText || options[i].text;
            if (text.indexOf('{activity_code}') === 0) {{
                targetValue = options[i].value;
                break;
            }}
        }}
        
        if (targetValue) {{
            select.val(targetValue);
            select.select2('data', {{id: targetValue, text: text}});
            select.trigger('change');
            return text;
        }}
        
        return null;
        """
        
        activity_name = driver.execute_script(set_value_script)
        
        if activity_name:
            print(f"    发现活动: {activity_name}")
            print(f"✓ 成功选择活动: {activity_name}")
            return activity_name
        else:
            print("    方法 2 也失败了")
            return "查無此活動"
    
    except Exception as e:
        print(f"✗ 错误: {e}")
        import traceback
        traceback.print_exc()
        return "查無此活動"

def write_to_excel(activity_name: str):
    """将活动名称写入 Excel B3"""
    if not os.path.exists(EXCEL_PATH):
        print(f"✗ 找不到 Excel 文件: {EXCEL_PATH}")
        return False
    
    try:
        print(f"\n[7] 打开 Excel 文件...")
        workbook = load_workbook(EXCEL_PATH)
        worksheet = workbook.active
        
        print(f"[8] 将 '{activity_name}' 写入 B3...")
        worksheet['B3'] = activity_name
        
        workbook.save(EXCEL_PATH)
        print(f"✓ 成功写入 B3: {activity_name}")
        return True
    
    except Exception as e:
        print(f"✗ 无法写入 Excel: {e}")
        return False

def main():
    """主函数"""
    # 检查 Excel 文件
    if not os.path.exists(EXCEL_PATH):
        print(f"✗ 找不到 Excel 文件: {EXCEL_PATH}")
        return
    
    # 读取活动代码
    print("[0] 读取 Excel 文件...")
    workbook = load_workbook(EXCEL_PATH)
    worksheet = workbook.active
    activity_code = worksheet['A2'].value
    
    if not activity_code:
        print("✗ Excel A2 为空，请先填入活动代码。")
        return
    
    print(f"✓ 从 A2 读取活动代码: {activity_code}")
    
    # 启动浏览器
    options = Options()
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    
    try:
        # 登入
        login_to_sms(driver)
        
        # 获取活动名称
        activity_name = get_activity_name(driver, str(activity_code))
        
        # 写入 Excel
        write_to_excel(activity_name)
        
        print("\n✓ 测试完成！")
    
    except Exception as e:
        print(f"\n✗ 错误: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
