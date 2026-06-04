#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试：查看实际获取的学生数据
"""

import sys
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

requests.packages.urllib3.disable_warnings()

print("=" * 100)
print("调试：查看实际获取的学生数据")
print("=" * 100)

# 读取 Excel
excel_path = "calligraphy.xlsx"
wb = load_workbook(excel_path, data_only=True)
ws = wb.active

# 从 Excel 读取所需班级
required_classes = set()
for row_num in range(5, ws.max_row + 1):
    class_name = ws.cell(row=row_num, column=2).value
    if class_name:
        required_classes.add(class_name)

wb.close()

print(f"\nExcel 中的班级: {required_classes}")

# 建立班级映射
class_name_to_id = {
    'J3B': '726',
    'S1B': '734',
}

required_class_ids = set()
for class_name in required_classes:
    if class_name in class_name_to_id:
        required_class_ids.add(class_name_to_id[class_name])

print(f"映射到的班级 ID: {required_class_ids}")

# 登录
session = requests.Session()
session.verify = False

LOGIN_URL = "http://sms.chhsban.edu.my/sms/index.php?r=site/login"
login_data = {
    'LoginForm[username]': 'schhs334',
    'LoginForm[password]': 'schhs334',
    'login-button': 'login'
}

session.get(LOGIN_URL, timeout=15)
session.post(LOGIN_URL, data=login_data, timeout=15, allow_redirects=True)

print("\n✓ 登录成功")

# 获取 item_id
ACTIVITY_PAGE = "http://sms.chhsban.edu.my/sms/index.php?r=transaction/studentPerformance/create"
resp = session.get(ACTIVITY_PAGE, timeout=15)
soup = BeautifulSoup(resp.text, 'html.parser')

select_element = soup.select_one('select#StudentPerformanceM_item_id')
item_id = ""

if select_element:
    for option in select_element.select('option[value]:not([value=""])'):
        option_text = option.get_text(strip=True)
        option_value = option.get('value')
        
        if 'ACA' in option_text and 'CMO207' in option_text:
            item_id = option_value
            print(f"\n✓ 找到活动: {option_text}")
            print(f"  item_id: {item_id}")
            break

# 对每个班级获取学生
print(f"\n获取学生数据:")
all_students_map = {}

for class_id in sorted(required_class_ids):
    print(f"\n班级 ID: {class_id}")
    
    ajax_url = "http://sms.chhsban.edu.my/sms/index.php"
    ajax_params = {
        'r': 'transaction/studentPerformance/update',
        'StudentPerformanceM[class_id]': class_id,
        'StudentPerformanceM[item_id]': item_id,
        'ajax': 'student-grid',
        'date': '2026-02-06',
        'item_id': item_id,
    }
    
    resp = session.get(ajax_url, params=ajax_params, timeout=15)
    soup = BeautifulSoup(resp.text, 'html.parser')
    
    links = soup.select('a[data-student_id]')
    
    print(f"  获取到 {len(links)} 个学生")
    
    all_students_map[class_id] = {}
    
    # 查找我们要的学生
    target_ids = ['24177', '23121', '23073']
    
    found_count = 0
    for link in links:
        student_no = link.get('data-student_no')
        
        if student_no in target_ids:
            internal_id = link.get('data-student_id')
            name = link.get('data-student_name')
            
            all_students_map[class_id][student_no] = {
                'internal_id': internal_id,
                'name': name,
            }
            
            print(f"    ✓ 找到: {student_no} - {name} (internal_id: {internal_id})")
            found_count += 1
    
    if found_count == 0:
        print(f"    ✗ 没有找到目标学生，显示前 5 个:")
        for i, link in enumerate(links[:5]):
            student_no = link.get('data-student_no')
            name = link.get('data-student_name')
            print(f"      [{i}] {student_no} - {name}")

print("\n" + "=" * 100)
print("完整的学生映射表:")
print("=" * 100)

for class_id, students in all_students_map.items():
    print(f"类 {class_id}: {len(students)} 个学生")
    for student_no, student_info in students.items():
        print(f"  {student_no}: {student_info['name']}")

# 测试匹配
print("\n" + "=" * 100)
print("测试学生匹配:")
print("=" * 100)

target_students = {
    '24177': 'J3B',
    '23121': 'S1B',
    '23073': 'S1B',
}

for student_id, expected_class in target_students.items():
    found = False
    for class_id, students in all_students_map.items():
        if student_id in students:
            print(f"✓ {student_id} 找到")
            found = True
            break
    
    if not found:
        print(f"✗ {student_id} 未找到")
