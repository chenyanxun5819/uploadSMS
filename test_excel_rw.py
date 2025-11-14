"""
简化版测试：读取 Excel A2 活动代码，模拟抓取结果，写入 B3
（用于验证 Excel 读写功能，暂不依赖网站登入）
"""

import os
from openpyxl import load_workbook

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Upload.xlsx")

def test_excel_read_write():
    """测试读取 Excel A2 并写入 B3"""
    
    if not os.path.exists(EXCEL_PATH):
        print(f"✗ 找不到 Excel 文件: {EXCEL_PATH}")
        return False
    
    try:
        print(f"[1] 打开 Excel 文件...")
        workbook = load_workbook(EXCEL_PATH)
        worksheet = workbook.active
        print(f"   ✓ 工作表: {worksheet.title}")
        
        # 读取 A2（活动代码）
        activity_code = worksheet['A2'].value
        print(f"\n[2] 读取 A2 活动代码...")
        print(f"   A2 = {activity_code}")
        
        if not activity_code:
            print("   ✗ A2 为空!")
            return False
        
        # 模拟抓取活动名称（这里先用 mock 数据代替）
        # 实际环境应该从网站获取
        mock_activity_name = f"活动: {activity_code}"
        print(f"\n[3] 模拟抓取活动名称...")
        print(f"   模拟结果: {mock_activity_name}")
        
        # 写入 B3
        print(f"\n[4] 写入 B3...")
        worksheet['B3'] = mock_activity_name
        print(f"   B3 = {mock_activity_name}")
        
        # 保存
        print(f"\n[5] 保存 Excel 文件...")
        workbook.save(EXCEL_PATH)
        print(f"   ✓ 已保存")
        
        # 验证
        print(f"\n[6] 验证写入是否成功...")
        workbook2 = load_workbook(EXCEL_PATH)
        worksheet2 = workbook2.active
        b3_value = worksheet2['B3'].value
        print(f"   B3 = {b3_value}")
        
        if b3_value == mock_activity_name:
            print(f"\n✓ 测试成功！Excel 读写功能正常。")
            return True
        else:
            print(f"\n✗ 验证失败！写入的值与读取的值不一致。")
            return False
    
    except Exception as e:
        print(f"✗ 错误: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_excel_read_write()
