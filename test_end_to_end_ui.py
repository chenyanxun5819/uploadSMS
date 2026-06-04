#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
最终的端到端测试 - 模拟 UI 流程
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / 'sms_app'))

from openpyxl import load_workbook
from core.sms_handler import SMSHandler

print("=" * 100)
print("最终端到端测试 - 模拟 UI 上传流程")
print("=" * 100)

# Step 1: 按 score_upload_page.py 的方式读取 Excel
print("\n[Step 1] 读取 Excel 文件...")

excel_path = Path("calligraphy.xlsx")

if not excel_path.exists():
    print(f"✗ 找不到 {excel_path}")
    sys.exit(1)

try:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    # 读取日期（B1）
    date_val = ws.cell(row=1, column=2).value
    if isinstance(date_val, str):
        date_str = date_val
    else:
        try:
            date_str = date_val.strftime('%Y-%m-%d')
        except:
            date_str = '2026-01-01'
    
    # 读取数据（从第5行开始）
    scores_data = []
    for row_num in range(5, ws.max_row + 1):
        cell_values = {
            'name': ws.cell(row=row_num, column=1).value,      # Column 1: 姓名
            'class': ws.cell(row=row_num, column=2).value,     # Column 2: 班级
            'student_id': str(ws.cell(row=row_num, column=3).value), # Column 3: 学号 -> 转为字符串！
            'remarks': ws.cell(row=row_num, column=4).value,   # Column 4: 奖项/备注
            'english_name': ws.cell(row=row_num, column=5).value, # Column 5: 英文名
        }
        
        # 如果至少有学号和班级，则视为有效行
        if cell_values['student_id'] and cell_values['class']:
            scores_data.append(cell_values)
    
    wb.close()
    
    print(f"✓ 读取到 {len(scores_data)} 条学生数据")
    print(f"✓ 上传日期: {date_str}")
    
    # 显示读取的数据
    print("\n读取的学生数据:")
    for i, student in enumerate(scores_data, 1):
        print(f"  [{i}] {student['class']:5} | {student['student_id']:10} | {student['name']:15} | {student['remarks']}")
    
except Exception as e:
    print(f"✗ 读取 Excel 失败: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# Step 2: 调用 SMS 处理器上传
print("\n[Step 2] 创建 SMS 处理器并上传...")

try:
    handler = SMSHandler()
    
    result = handler.upload_student_scores(
        username='schhs334',
        password='schhs334',
        scores_data=scores_data,
        date=date_str,
        activity_code='ACA CMO207'
    )
    
    # 显示结果
    print("\n" + "=" * 100)
    print("上传结果")
    print("=" * 100)
    
    if result['success']:
        print(f"\n✓ 状态: 成功")
        print(f"✓ 消息: {result['message']}")
        print(f"✓ 上传人数: {result['total']}")
        if result.get('errors'):
            print(f"\n⚠️  未上传的学生:")
            for error in result['errors']:
                print(f"   - {error}")
    else:
        print(f"\n✗ 状态: 失败")
        print(f"✗ 消息: {result['message']}")
        if result.get('errors'):
            print(f"\n错误信息:")
            for error in result['errors']:
                print(f"   - {error}")
    
    print("\n" + "=" * 100)
    
    # 返回状态码
    sys.exit(0 if result['success'] else 1)

except Exception as e:
    print(f"✗ 上传失败: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
