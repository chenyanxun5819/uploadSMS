#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
完整工作流程：
1. 登入 SMS
2. 進入活動頁面
3. 填寫日期（A1）與活動代碼（A2）
4. 點擊學生名單按鈕
5. 按班級逐一查詢學生，抓取英文名
6. 將英文名寫入 Excel E 欄

用法：
  $env:SMS_USERNAME = "帳號"; $env:SMS_PASSWORD = "密碼"; python fill_english_names_v2.py

環境變數：
  SMS_USERNAME, SMS_PASSWORD
可選：
  HEADLESS=1  # 無頭模式
"""
import os
import json
import time
import sys
from collections import defaultdict
from datetime import datetime
from typing import Optional, Dict, List

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager


SMS_LOGIN = "http://sms.chhsban.edu.my/sms/index.php?r=site/login"
SMS_ACTIVITY_PAGE = "http://sms.chhsban.edu.my/sms/index.php?r=transaction/studentPerformance/create"
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Upload.xlsx")
SETTING_FILE = os.path.join(os.path.dirname(__file__), "setting.json")


def load_field_mapping():
    """從 Excel 第 4 行讀取欄位名稱，返回 {field_name: column_index}"""
    try:
        excel_file = EXCEL_FILE
        if not os.path.exists(excel_file):
            print(f'⚠ 找不到 Excel，嘗試從 setting.json 讀取...')
            raise FileNotFoundError('Excel not found')
        
        # 優先從 Excel 第 4 行讀取標題
        wb = load_workbook(excel_file, data_only=False)
        ws = wb.active
        
        mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=4, column=col_idx).value
            if cell_value:
                field_name = str(cell_value).strip().lower()
                mapping[field_name] = col_idx
                print(f'    [欄位] 第 {col_idx} 列: {field_name}')
        
        if mapping:
            print(f'  ✓ 從 Excel 第 4 行讀取欄位對應: {mapping}')
            wb.close()
            return mapping
        else:
            print(f'⚠ Excel 第 4 行為空，嘗試從 setting.json 讀取...')
            wb.close()
            raise ValueError('Row 4 is empty')
    
    except Exception as e:
        print(f'  ⚠ 無法從 Excel 讀取欄位，轉用 setting.json: {e}')
        try:
            if os.path.exists(SETTING_FILE):
                with open(SETTING_FILE, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    fields = data.get('student_fields', [])
                    mapping = {}
                    for i, field in enumerate(fields):
                        mapping[field.lower()] = i + 1  # Excel 欄位從 1 開始（A=1, B=2...）
                    print(f'  ✓ 從 setting.json 讀取欄位對應: {mapping}')
                    return mapping
        except Exception as e2:
            print(f'  ⚠ 無法讀取 setting.json: {e2}')
    
    # 預設欄位順序（若都讀取失敗）
    print(f'  ⚠ 使用預設欄位順序')
    return {
        'class': 1,
        'studentid': 2,
        'name': 3,
        'award': 4
    }


def setup_driver(headless: bool = False):
    """初始化 Selenium WebDriver"""
    options = Options()
    if headless:
        options.add_argument('--headless=new')
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), 
        options=options
    )
    driver.set_window_size(1200, 900)
    return driver


def login(driver, username: str, password: str, timeout: int = 10) -> bool:
    """登入 SMS"""
    print('[1/6] 連接登入頁面...')
    driver.get(SMS_LOGIN)
    try:
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, 'LoginForm_username'))
        )
        print('  填入帳號與密碼...')
        driver.find_element(By.ID, 'LoginForm_username').send_keys(username)
        driver.find_element(By.ID, 'LoginForm_password').send_keys(password)

        print('  提交表單...')
        driver.find_element(By.XPATH, "//button[@type='submit']").click()

        print('  等待登入完成...')
        WebDriverWait(driver, timeout).until(
            lambda d: 'login' not in d.current_url.lower()
        )
        time.sleep(1)
        print('✓ 已登入')
        return True
    except Exception as e:
        print(f'✗ 登入失敗: {e}')
        return False


def fill_date_and_activity(driver, date_str: str, activity_code: str, timeout: int = 8) -> bool:
    """填寫日期與活動代碼"""
    print('[2/6] 進入活動頁面並填寫基本資料...')
    driver.get(SMS_ACTIVITY_PAGE)
    time.sleep(2)

    try:
        # 填寫日期 (A1 → ID: StudentPerformanceM_date)
        print(f'  填寫日期: {date_str}')
        date_field = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, 'StudentPerformanceM_date'))
        )
        date_field.clear()
        date_field.send_keys(date_str)
        time.sleep(0.5)

        # 填寫活動代碼 (A2 → Select2: StudentPerformanceM_item_id)
        print(f'  選擇活動: {activity_code}')
        if not select_activity(driver, activity_code):
            print('⚠ 活動選擇失敗，但嘗試繼續')

        time.sleep(1)
        print('✓ 基本資料已填寫')
        return True

    except Exception as e:
        print(f'✗ 填寫失敗: {e}')
        return False


def select_activity(driver, activity_code: str) -> bool:
    """透過 Select2 選擇活動"""
    try:
        # 嘗試方法 1: Select2 搜尋
        script = f"""
        var select = $('#StudentPerformanceM_item_id');
        select.select2('open');
        var searchInput = document.querySelector('.select2-input');
        if (searchInput) {{
            searchInput.value = '{activity_code}';
            searchInput.dispatchEvent(new Event('input', {{bubbles: true}}));
            searchInput.dispatchEvent(new Event('change', {{bubbles: true}}));
        }}
        return true;
        """
        driver.execute_script(script)
        time.sleep(1)

        # 嘗試點擊搜尋結果
        try:
            result = driver.find_element(By.CSS_SELECTOR, '.select2-result-label')
            result.click()
            time.sleep(1)
            return True
        except:
            pass

        # 方法 2: 直接設定 value
        script2 = f"""
        var select = $('#StudentPerformanceM_item_id');
        var options = select.find('option');
        for (var i = 0; i < options.length; i++) {{
            var text = options[i].text || options[i].innerText || '';
            if (text.indexOf('{activity_code}') === 0) {{
                select.val(options[i].value);
                select.select2('data', {{id: options[i].value, text: text}});
                select.trigger('change');
                return true;
            }}
        }}
        return false;
        """
        result = driver.execute_script(script2)
        if result:
            time.sleep(1)
            return True

        print('⚠ 活動搜尋失敗')
        return False
    except Exception as e:
        print(f'⚠ Select2 操作異常: {e}')
        return False


def click_student_list_button(driver, timeout: int = 8) -> bool:
    """點擊學生名單按鈕"""
    print('[3/6] 點擊學生名單按鈕...')
    try:
        btn = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.ID, 'yw4'))
        )
        btn.click()
        print('  等待班級選擇下拉出現...')
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, 'class_id'))
        )
        time.sleep(1)
        print('✓ 學生名單已打開')
        return True
    except Exception as e:
        print(f'✗ 點擊失敗: {e}')
        return False


def select_class(driver, class_short: str, timeout: int = 8) -> bool:
    """選擇班級"""
    try:
        sel = WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.ID, 'class_id'))
        )
        select = Select(sel)

        # 匹配班級簡寫（括號內或尾部）
        matched_value = None
        for opt in sel.find_elements(By.TAG_NAME, 'option'):
            text = opt.text.strip()
            if f'({class_short})' in text or text.endswith(class_short):
                matched_value = opt.get_attribute('value')
                break

        if not matched_value:
            print(f'⚠ 找不到班級: {class_short}')
            return False

        select.select_by_value(matched_value)
        print(f'  已選擇班級: {class_short}')

        # 等待表格載入
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'table.table tbody tr'))
        )
        time.sleep(0.8)
        return True

    except Exception as e:
        print(f'✗ 選班失敗: {e}')
        return False


def find_student_in_table(driver, student_id: str) -> Optional[Dict]:
    """在表格中查找學生並返回資料及操作按鈕"""
    try:
        rows = driver.find_elements(By.CSS_SELECTOR, 'table.table tbody tr')
        for r in rows:
            cols = r.find_elements(By.TAG_NAME, 'td')
            if not cols or len(cols) < 5:
                continue
            stu_no = cols[0].text.strip()
            if stu_no == str(student_id):
                # cols[0]=學號, cols[1]=英文名, cols[2]=中文名, cols[3]=班級, cols[4]=操作（含按鈕）
                name_en = cols[1].text.strip()
                
                # 尋找該列的「添加」按鈕
                try:
                    btn = r.find_element(By.CSS_SELECTOR, 'a.btn[onclick*="addToEkstra"]')
                    return {
                        'student_no': stu_no,
                        'name_en': name_en,
                        'button': btn
                    }
                except:
                    print(f'    ⚠ 找到學生 {student_id}，但找不到操作按鈕')
                    return None
        return None
    except Exception as e:
        print(f'✗ 表格查詢錯誤: {e}')
        return None


def main():
    if not os.path.exists(EXCEL_FILE):
        print(f'✗ 找不到 Excel: {EXCEL_FILE}')
        return

    # 帳號密碼（寫死）
    username = 'schhs334'
    password = 'schhs334'

    headless = bool(os.getenv('HEADLESS'))

    # 讀取欄位對應
    print('[0/6] 讀取設定和 Excel 資料...')
    field_map = load_field_mapping()
    
    # 讀取 Excel
    print('[0/6] 讀取 Excel 資料...')
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # 讀取事件資訊
    date_val = ws['A1'].value
    activity_code = ws['A2'].value

    if not date_val or not activity_code:
        print('✗ Excel A1 或 A2 為空')
        return

    # 轉換日期格式
    try:
        if isinstance(date_val, str):
            date_str = date_val
        else:
            # 若為 datetime 物件，轉為 yyyy-MM-dd
            date_str = date_val.strftime('%Y-%m-%d')
    except Exception as e:
        print(f'✗ 日期轉換失敗: {e}')
        return

    print(f'  日期: {date_str}, 活動: {activity_code}')

    # 讀取學生清單（Row 5+）
    # Row 4 是標題列
    students_by_class = defaultdict(list)  # class_short -> [(row_idx, student_id), ...]
    seen_pairs = set()  # 用來去重：(class, student_id)

    class_col = field_map.get('class', field_map.get('Class', 1))
    student_id_col = field_map.get('studentid', field_map.get('studentId', field_map.get('student_id', 2)))
    
    print(f'  班級欄: {class_col}, 學號欄: {student_id_col}')

    for row_idx in range(5, ws.max_row + 1):
        class_val = ws.cell(row=row_idx, column=class_col).value
        id_val = ws.cell(row=row_idx, column=student_id_col).value

        if not id_val:
            continue

        class_short = str(class_val).strip() if class_val else ''
        if not class_short:
            continue

        student_id = str(id_val).strip()
        pair = (class_short, student_id)
        
        # 去重：只保留第一次出現
        if pair in seen_pairs:
            print(f'  [去重] 跳過重複的 {class_short} - {student_id}（第 {row_idx} 行）')
            continue
        
        seen_pairs.add(pair)
        students_by_class[class_short].append((row_idx, student_id))

    if not students_by_class:
        print('✗ Excel 中無學生資料')
        return

    print(f'✓ 讀取到 {len(students_by_class)} 個班級，共 {sum(len(v) for v in students_by_class.values())} 位學生')

    # 初始化瀏覽器
    driver = setup_driver(headless=headless)
    found_count = 0
    missing_count = 0

    try:
        # 登入
        if not login(driver, username, password):
            return

        # 填寫日期與活動
        if not fill_date_and_activity(driver, date_str, str(activity_code)):
            return

        # 點擊學生名單
        if not click_student_list_button(driver):
            return

        # [4/6] 逐班級查詢學生
        print('[4/6] 逐班級查詢學生英文名...')
        class_count = 0

        for class_short, entries in sorted(students_by_class.items()):
            class_count += 1
            print(f'\n  [{class_count}/{len(students_by_class)}] 班級: {class_short}（{len(entries)} 位）')

            # 選擇班級
            if not select_class(driver, class_short):
                print(f'    跳過班級 {class_short}')
                missing_count += len(entries)
                continue

            # 逐個查找該班學生
            for row_idx, student_id in entries:
                result = find_student_in_table(driver, student_id)
                if result:
                    name_en = result['name_en']
                    btn = result['button']
                    print(f'    ✓ {student_id} → {name_en}')
                    
                    # 點擊「添加」按鈕
                    try:
                        btn.click()
                        time.sleep(0.8)  # 延長延遲，讓 JavaScript 完成
                        print(f'      已添加到名單')
                        
                        # ...不再寫入英文名到 Excel...
                        found_count += 1
                    except Exception as e:
                        print(f'      ⚠ 點擊按鈕失敗: {e}')
                        missing_count += 1
                else:
                    print(f'    ⚠ {student_id} 未找到')
                    missing_count += 1

        # [5/6] 關閉 Modal，返回上一頁
        print(f'\n[5/6] 關閉學生名單...')
        try:
            # 尋找 Modal 關閉按鈕（class="close"）
            close_btn = driver.find_element(By.CSS_SELECTOR, '#studentModal a.close')
            close_btn.click()
            time.sleep(1)
            print('✓ Modal 已關閉，回到上一頁')
        except Exception as e:
            print(f'⚠ 關閉 Modal 失敗，嘗試用 Escape 鍵: {e}')
            try:
                driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                time.sleep(1)
            except:
                pass

        # [5.5/6] 步驟 9.5 和 9.6：為每位學生填寫「奪勵分數類型」和「備註」
        print('\n[5.5/6] 填寫奪勵分數類型和備註...')
        time.sleep(1)  # 等待頁面加載完成
        
        try:
            # 建立一個集合，只包含我們剛才添加的學生學號
            added_students_set = set()
            for class_short, entries in students_by_class.items():
                for _, student_id in entries:
                    added_students_set.add(str(student_id))
            
            print(f'  本次添加的學生: {added_students_set}')
            
            # 抓取頁面表格中的所有行
            table_rows = driver.find_elements(By.CSS_SELECTOR, 'table.table tbody tr')
            print(f'  頁面表格總共 {len(table_rows)} 行（包含舊資料）')
            
            processed_count = 0
            for row in table_rows:
                try:
                    cols = row.find_elements(By.TAG_NAME, 'td')
                    if not cols or len(cols) < 5:
                        continue
                    
                    # 從表格第 1 欄（cols[0]）抓取學號
                    student_no_from_page = cols[0].text.strip()
                    
                    # 只處理我們本次添加的學生
                    if student_no_from_page not in added_students_set:
                        continue
                    
                    processed_count += 1
                    print(f'  處理學號: {student_no_from_page}')
                    
                    # 在 Excel 中查找相同學號的行
                    excel_remark = None
                    for row_idx in range(5, ws.max_row + 1):
                        excel_student_id = ws.cell(row=row_idx, column=student_id_col).value
                        if str(excel_student_id).strip() == student_no_from_page:
                            # 備註欄位是 'award' 欄位
                            award_col = field_map.get('award', field_map.get('Award', 4))
                            excel_remark = ws.cell(row=row_idx, column=award_col).value or ""
                            print(f'    ✓ 匹配到 Excel 行 {row_idx}，備註: {excel_remark}')
                            break
                    
                    if excel_remark is None:
                        print(f'    ⚠ 在 Excel 中未找到學號 {student_no_from_page}')
                        continue
                    
                    # 從表格行的 class 屬性提取內部 ID
                    tr_class = row.get_attribute('class')
                    internal_id = tr_class.strip() if tr_class else None
                    
                    if not internal_id:
                        print(f'    ⚠ 無法獲取內部 ID')
                        continue
                    
                    # 9.5: 選擇「校外學藝」
                    try:
                        select_elem = driver.find_element(
                            By.CSS_SELECTOR,
                            f'select[name="StudentPerformanceM[inputperformance][{internal_id}][type_of_bonus]"]'
                        )
                        select_obj = Select(select_elem)
                        select_obj.select_by_value("1")  # "1" = 校外學藝
                        print(f'    ✓ 已選擇「校外學藝」')
                    except Exception as e:
                        print(f'    ⚠ 選擇「校外學藝」失敗: {e}')
                    
                    # 9.6: 填寫備註
                    try:
                        textarea_elem = driver.find_element(
                            By.ID,
                            f'StudentPerformanceM_inputperformance_{internal_id}_remark'
                        )
                        textarea_elem.clear()
                        textarea_elem.send_keys(str(excel_remark))
                        print(f'    ✓ 已填寫備註: {excel_remark}')
                    except Exception as e:
                        print(f'    ⚠ 填寫備註失敗: {e}')
                
                except Exception as e:
                    print(f'  ⚠ 處理行時出錯: {e}')
            
            print(f'✓ 已完成填寫 {processed_count} 位學生的奪勵分數類型和備註')
        
        except Exception as e:
            print(f'⚠ 步驟 9.5/9.6 出錯: {e}')

        # [6/6] 點擊「創建」按鈕
        print('\n[6/6] 提交表單...')
        try:
            submit_btn = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.ID, 'yw7'))
            )
            submit_btn.click()
            print('✓ 已提交')
            time.sleep(2)
        except Exception as e:
            print(f'⚠ 提交失敗: {e}')

        # 統計
        print(f'\n完成')
        print(f'  成功填寫並提交: {found_count}')
        print(f'  未找到: {missing_count}')
        print(f'\n✓ 流程結束，請在瀏覽器中檢查結果')
        
        # 保留瀏覽器窗口
        print('\n浏览器已保持打开。按 Enter 退出...')
        input()

    except KeyboardInterrupt:
        print('\n✗ 中斷執行')
    except Exception as e:
        print(f'\n✗ 發生錯誤: {e}')
        import traceback
        traceback.print_exc()
    finally:
        try:
            driver.quit()
        except:
            pass


if __name__ == '__main__':
    main()
