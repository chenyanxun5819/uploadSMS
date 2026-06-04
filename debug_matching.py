#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
调试脚本 - 检查学生匹配是否正确
"""

import sys
from pathlib import Path
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests

# 禁用 SSL 警告
requests.packages.urllib3.disable_warnings()

print("=" * 80)
print("检查 Excel 数据和 SMS 系统的学生匹配")
print("=" * 80)

# 1. 读取 Excel 数据
print("\n[1/3] 读取 Excel 数据...")
excel_path = Path(__file__).parent / "calligraphy.xlsx"

if excel_path.exists():
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # 获取项目代码（第2行）
    project_code = ws.cell(row=2, column=1).value
    print(f"  项目代码: {project_code}")
    
    scores_data = []
    for row_idx in range(5, min(10, ws.max_row + 1)):  # 只取前5个学生用于测试
        name = ws.cell(row=row_idx, column=1).value
        class_name = ws.cell(row=row_idx, column=2).value
        student_id = ws.cell(row=row_idx, column=3).value
        award = ws.cell(row=row_idx, column=4).value
        
        if not student_id or not class_name:
            continue
        
        scores_data.append({
            'name': name,
            'class': class_name,
            'student_id': str(student_id),
            'remarks': str(award) if award else ''
        })
    
    print(f"  读取 {len(scores_data)} 条记录（仅用于测试）:")
    for score in scores_data:
        print(f"    - Class: {score['class']:10} | StudentID: {score['student_id']:10} | Name: {score['name']}")
else:
    print(f"  ✗ 文件未找到: {excel_path}")
    sys.exit(1)

# 2. 从 SMS 系统提取学生数据（需要登录）
print("\n[2/3] 从 SMS 系统提取学生数据...")
print("  (需要登录)")

username = input("  请输入用户名: ").strip()
password = input("  请输入密码: ").strip()

if not username or not password:
    print("  ✗ 用户名或密码为空")
    sys.exit(1)

try:
    session = requests.Session()
    session.verify = False
    
    # 登录
    print("  正在登录...")
    LOGIN_URL = "http://sms.chhsban.edu.my/sms/index.php?r=site/login"
    login_data = {
        'LoginForm[username]': username,
        'LoginForm[password]': password,
        'login-button': 'login'
    }
    
    session.get(LOGIN_URL, timeout=15)
    resp = session.post(LOGIN_URL, data=login_data, timeout=15, allow_redirects=True)
    
    if 'login' in resp.url.lower():
        print("  ✗ 登录失败")
        sys.exit(1)
    
    print("  ✓ 登录成功")
    
    # 获取活动页面
    print("  正在获取学生数据...")
    ACTIVITY_PAGE = "http://sms.chhsban.edu.my/sms/index.php?r=transaction/studentPerformance/create"
    resp = session.get(ACTIVITY_PAGE, timeout=15)
    soup = BeautifulSoup(resp.text, 'html.parser')
    
    # 提取所有学生
    all_students_map = {}
    for link in soup.select('a[data-student_id]'):
        student = {
            'internal_id': link.get('data-student_id'),
            'student_no': link.get('data-student_no'),
            'name': link.get('data-student_name'),
            'class_id': link.get('data-class_id'),
            'class_name': link.get('data-class_name'),
        }
        
        if student['internal_id'] and student['student_no']:
            if student['class_id'] not in all_students_map:
                all_students_map[student['class_id']] = {}
            all_students_map[student['class_id']][student['student_no']] = student
    
    print(f"  ✓ 从 SMS 系统获取 {sum(len(v) for v in all_students_map.values())} 条学生记录")
    
    # 3. 尝试匹配
    print("\n[3/3] 尝试匹配 Excel 和 SMS 系统中的学生...")
    
    matched = 0
    not_matched = []
    
    for score_item in scores_data:
        student_id = score_item['student_id']
        class_name = score_item.get('class', '')
        
        found = False
        for class_id, students_in_class in all_students_map.items():
            if student_id in students_in_class:
                sms_student = students_in_class[student_id]
                print(f"  ✓ {class_name:10} {student_id:10} -> internal_id: {sms_student['internal_id']:10} class_id: {sms_student['class_id']}")
                matched += 1
                found = True
                break
        
        if not found:
            print(f"  ✗ {class_name:10} {student_id:10} -> 未找到")
            not_matched.append(f"{class_name} {student_id}")
    
    print(f"\n  === 匹配结果 ===")
    print(f"  成功匹配: {matched}/{len(scores_data)}")
    print(f"  未找到: {len(not_matched)}/{len(scores_data)}")
    
    if not_matched:
        print(f"\n  未找到的学生:")
        for student in not_matched:
            print(f"    - {student}")
    
except Exception as e:
    print(f"  ✗ 错误: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

print("\n" + "=" * 80)
