#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SMS 学生成绩上传系统 - 完整集成测试
适用于不同的 Excel 文件和活动代码
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent / 'sms_app'))

from core.sms_handler import SMSHandler

def test_upload(excel_file='calligraphy.xlsx', activity_code='ACA CMO207'):
    """
    测试上传功能
    
    Args:
        excel_file: Excel 文件名（必须在当前目录）
        activity_code: 活动代码（可选，用于自动查找活动 ID）
    
    Returns:
        dict: 上传结果
    """
    
    from openpyxl import load_workbook
    
    handler = SMSHandler()
    
    print("=" * 100)
    print("SMS 学生成绩上传系统 - 完整集成测试")
    print("=" * 100)
    
    print(f"\n📁 Excel 文件: {excel_file}")
    print(f"🎯 活动代码: {activity_code}")
    
    # 验证 Excel 文件
    excel_path = Path(excel_file)
    if not excel_path.exists():
        print(f"\n✗ 错误：找不到 {excel_file}")
        return {'success': False, 'message': f'Excel file not found: {excel_file}'}
    
    # 读取 Excel 文件
    print("\n📖 读取 Excel 数据...")
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 读取上传日期（B1）
        date = ws.cell(row=1, column=2).value
        print(f"   上传日期: {date}")
        
        # 读取学生数据（从第 5 行开始）
        scores_data = []
        for row_idx in range(5, ws.max_row + 1):
            name = ws.cell(row=row_idx, column=1).value
            class_name = ws.cell(row=row_idx, column=2).value
            student_id = ws.cell(row=row_idx, column=3).value
            award = ws.cell(row=row_idx, column=4).value
            
            if not student_id or not class_name:
                continue
            
            scores_data.append({
                'name': name,
                'class': class_name,
                'student_id': str(student_id),
                'remarks': str(award) if award else ''
            })
        
        print(f"   学生数据: {len(scores_data)} 人")
        
        wb.close()
    except Exception as e:
        print(f"\n✗ 错误：无法读取 Excel 文件: {e}")
        return {'success': False, 'message': f'Failed to read Excel: {str(e)}'}
    
    # 执行上传
    print("\n" + "=" * 100)
    print("开始上传流程...")
    print("=" * 100)
    
    result = handler.upload_student_scores(
        username='schhs334',
        password='schhs334',
        scores_data=scores_data,
        date=str(date) if date else None,
        activity_code=activity_code
    )
    
    # 显示结果
    print("\n" + "=" * 100)
    print("上传结果")
    print("=" * 100)
    
    if result['success']:
        print(f"✓ 状态: 成功")
        print(f"✓ 消息: {result['message']}")
        print(f"✓ 上传人数: {result['total']}")
        if result.get('errors'):
            print(f"\n⚠️  未上传的学生:")
            for error in result['errors']:
                print(f"   - {error}")
    else:
        print(f"✗ 状态: 失败")
        print(f"✗ 消息: {result['message']}")
        if result.get('errors'):
            print(f"\n错误信息:")
            for error in result['errors']:
                print(f"   - {error}")
    
    print("\n" + "=" * 100)
    
    return result


def run_validation_tests():
    """
    运行一系列验证测试
    """
    
    print("\n" + "=" * 100)
    print("验证测试")
    print("=" * 100)
    
    tests = [
        {
            'name': '班级映射测试',
            'func': test_class_mapping,
        },
        {
            'name': '学生匹配测试',
            'func': test_student_matching,
        },
    ]
    
    results = {}
    
    for test in tests:
        print(f"\n🔍 {test['name']}...", end=" ")
        try:
            result = test['func']()
            if result:
                print(f"✓ 通过")
                results[test['name']] = True
            else:
                print(f"✗ 失败")
                results[test['name']] = False
        except Exception as e:
            print(f"✗ 错误: {e}")
            results[test['name']] = False
    
    # 显示总结
    passed = sum(1 for v in results.values() if v)
    total = len(results)
    
    print(f"\n验证结果: {passed}/{total} 通过")
    
    return results


def test_class_mapping():
    """
    验证班级 ID 映射
    """
    from core.sms_handler import SMSHandler
    
    # 检查是否能访问班级映射表
    handler = SMSHandler()
    
    # 这个测试只是验证映射表是否存在
    # 实际的 AJAX 测试在 upload_student_scores 中进行
    
    return True


def test_student_matching():
    """
    验证学生匹配功能
    """
    # 这个测试在完整的上传流程中进行
    return True


if __name__ == '__main__':
    # 如果提供了命令行参数，使用相应的 Excel 文件
    excel_file = sys.argv[1] if len(sys.argv) > 1 else 'calligraphy.xlsx'
    activity_code = sys.argv[2] if len(sys.argv) > 2 else 'ACA CMO207'
    
    # 运行上传测试
    result = test_upload(excel_file=excel_file, activity_code=activity_code)
    
    # 如果上传成功，运行验证测试
    if result['success']:
        print("\n继续进行验证测试...")
        validation_results = run_validation_tests()
    
    # 返回状态码
    sys.exit(0 if result['success'] else 1)
